import unicodedata
from typing import List, Optional
from database.db import get_connection
from models.modelos import Alimento, Cardapio, Refeicao, ItemRefeicao


def _sem_acento(texto: str) -> str:
    """Remove acentos e converte para minúsculas — usada como função SQL."""
    if not texto:
        return ""
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii").lower()


# ── Alimentos ─────────────────────────────────────────────────────────────────

def buscar_alimentos(termo: str = "", grupo: str = "") -> List[Alimento]:
    conn = get_connection()
    conn.create_function("sem_acento", 1, _sem_acento)
    cursor = conn.cursor()
    query = "SELECT * FROM alimentos WHERE 1=1"
    params = []
    if termo:
        query += " AND sem_acento(descricao) LIKE sem_acento(?)"
        params.append(f"%{termo}%")
    if grupo:
        query += " AND grupo = ?"
        params.append(grupo)
    query += " ORDER BY grupo, descricao"
    rows = cursor.execute(query, params).fetchall()
    conn.close()
    return [_row_to_alimento(r) for r in rows]


def listar_grupos() -> List[str]:
    """Retorna lista de todos os grupos (tabela grupos + grupos em alimentos)."""
    conn = get_connection()
    
    # Grupos da tabela grupos
    grupos_tabela = conn.execute("SELECT DISTINCT nome FROM grupos ORDER BY nome").fetchall()
    grupos_set = set(r["nome"] for r in grupos_tabela)
    
    # Grupos que existem em alimentos
    grupos_alimentos = conn.execute("SELECT DISTINCT grupo FROM alimentos ORDER BY grupo").fetchall()
    for r in grupos_alimentos:
        grupos_set.add(r["grupo"])
    
    conn.close()
    return sorted(list(grupos_set))


def buscar_alimento_por_id(alimento_id: int) -> Optional[Alimento]:
    conn = get_connection()
    row = conn.execute("SELECT * FROM alimentos WHERE id = ?", (alimento_id,)).fetchone()
    conn.close()
    return _row_to_alimento(row) if row else None


def _row_to_alimento(row) -> Alimento:
    return Alimento(
        id=row["id"],
        grupo=row["grupo"],
        descricao=row["descricao"],
        calorias=row["calorias"],
        proteinas=row["proteinas"],
        lipideos=row["lipideos"],
        carboidratos=row["carboidratos"],
        fonte=row["fonte"],
    )


# ── Cardápios ─────────────────────────────────────────────────────────────────

def listar_cardapios() -> List[dict]:
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, nome, kcal_total, data_criacao FROM cardapios ORDER BY data_criacao DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def carregar_cardapio(cardapio_id: int) -> Optional[Cardapio]:
    conn = get_connection()

    row = conn.execute("SELECT * FROM cardapios WHERE id = ?", (cardapio_id,)).fetchone()
    if not row:
        conn.close()
        return None

    cardapio = Cardapio(
        id=row["id"],
        nome=row["nome"],
        kcal_total=row["kcal_total"],
        data_criacao=row["data_criacao"],
    )

    ref_rows = conn.execute(
        "SELECT * FROM refeicoes WHERE cardapio_id = ? ORDER BY ordem", (cardapio_id,)
    ).fetchall()

    for ref_row in ref_rows:
        refeicao = Refeicao(
            id=ref_row["id"],
            nome=ref_row["nome"],
            percentual=ref_row["percentual"],
            kcal_diaria=cardapio.kcal_total,
            ordem=ref_row["ordem"],
        )

        item_rows = conn.execute("""
            SELECT ra.quantidade_g,
                   a.id, a.grupo, a.descricao, a.calorias,
                   a.proteinas, a.lipideos, a.carboidratos, a.fonte
            FROM refeicao_alimentos ra
            JOIN alimentos a ON a.id = ra.alimento_id
            WHERE ra.refeicao_id = ?
        """, (refeicao.id,)).fetchall()

        for item_row in item_rows:
            alimento = Alimento(
                id=item_row["id"],
                grupo=item_row["grupo"],
                descricao=item_row["descricao"],
                calorias=item_row["calorias"],
                proteinas=item_row["proteinas"],
                lipideos=item_row["lipideos"],
                carboidratos=item_row["carboidratos"],
                fonte=item_row["fonte"],
            )
            refeicao.itens.append(ItemRefeicao(
                alimento=alimento,
                quantidade_g=item_row["quantidade_g"],
            ))

        custom_rows = conn.execute(
            "SELECT * FROM refeicao_alimentos_custom WHERE refeicao_id = ?",
            (refeicao.id,),
        ).fetchall()
        for crow in custom_rows:
            alimento_custom = Alimento(
                id=None,
                grupo="",
                descricao=crow["descricao"],
                calorias=crow["calorias"],
                proteinas=crow["proteinas"],
                lipideos=crow["lipideos"],
                carboidratos=crow["carboidratos"],
                fonte="custom",
            )
            refeicao.itens.append(ItemRefeicao(
                id=crow["id"],
                alimento=alimento_custom,
                quantidade_g=crow["quantidade_g"],
            ))

        cardapio.refeicoes.append(refeicao)

    conn.close()
    return cardapio


def salvar_cardapio(cardapio: Cardapio) -> int:
    """Insere ou atualiza o cardápio. Retorna o id."""
    conn = get_connection()
    cursor = conn.cursor()

    if cardapio.id is None:
        cursor.execute(
            "INSERT INTO cardapios (nome, kcal_total) VALUES (?, ?)",
            (cardapio.nome, cardapio.kcal_total),
        )
        cardapio.id = cursor.lastrowid
    else:
        cursor.execute(
            "UPDATE cardapios SET nome = ?, kcal_total = ? WHERE id = ?",
            (cardapio.nome, cardapio.kcal_total, cardapio.id),
        )
        cursor.execute("DELETE FROM refeicoes WHERE cardapio_id = ?", (cardapio.id,))

    for ordem, refeicao in enumerate(cardapio.refeicoes):
        cursor.execute(
            "INSERT INTO refeicoes (cardapio_id, nome, percentual, ordem) VALUES (?, ?, ?, ?)",
            (cardapio.id, refeicao.nome, refeicao.percentual, ordem),
        )
        refeicao_id = cursor.lastrowid

        for item in refeicao.itens:
            if item.alimento.fonte == 'custom':
                cursor.execute(
                    """INSERT INTO refeicao_alimentos_custom
                       (refeicao_id, descricao, quantidade_g, calorias, proteinas, lipideos, carboidratos)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (refeicao_id, item.alimento.descricao, item.quantidade_g,
                     item.alimento.calorias, item.alimento.proteinas,
                     item.alimento.lipideos, item.alimento.carboidratos),
                )
            else:
                cursor.execute(
                    "INSERT INTO refeicao_alimentos (refeicao_id, alimento_id, quantidade_g) VALUES (?, ?, ?)",
                    (refeicao_id, item.alimento.id, item.quantidade_g),
                )

    conn.commit()
    conn.close()
    return cardapio.id


def excluir_cardapio(cardapio_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM cardapios WHERE id = ?", (cardapio_id,))
    conn.commit()
    conn.close()


# ── Gerenciamento de Alimentos e Grupos ──────────────────────────────────────

def adicionar_alimento(grupo: str, descricao: str, calorias: float, 
                       proteinas: float, lipideos: float, carboidratos: float) -> int:
    """Adiciona um novo alimento com fonte 'usuario'. Retorna o id."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO alimentos (grupo, descricao, calorias, proteinas, lipideos, carboidratos, fonte)
        VALUES (?, ?, ?, ?, ?, ?, 'usuario')
    """, (grupo, descricao, calorias, proteinas, lipideos, carboidratos))
    alimento_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return alimento_id


def editar_alimento(alimento_id: int, grupo: str, descricao: str, 
                    calorias: float, proteinas: float, lipideos: float, carboidratos: float):
    """Edita um alimento existente (apenas se fonte = 'usuario')."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Verifica se é alimento do usuário
    row = cursor.execute("SELECT fonte FROM alimentos WHERE id = ?", (alimento_id,)).fetchone()
    if not row or row["fonte"] != "usuario":
        conn.close()
        raise ValueError("Apenas alimentos criados pelo usuário podem ser editados.")
    
    cursor.execute("""
        UPDATE alimentos 
        SET grupo = ?, descricao = ?, calorias = ?, proteinas = ?, lipideos = ?, carboidratos = ?
        WHERE id = ?
    """, (grupo, descricao, calorias, proteinas, lipideos, carboidratos, alimento_id))
    
    conn.commit()
    conn.close()


def excluir_alimento(alimento_id: int):
    """Exclui um alimento (apenas se fonte = 'usuario')."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Verifica se é alimento do usuário
    row = cursor.execute("SELECT fonte FROM alimentos WHERE id = ?", (alimento_id,)).fetchone()
    if not row or row["fonte"] != "usuario":
        conn.close()
        raise ValueError("Apenas alimentos criados pelo usuário podem ser excluídos.")
    
    cursor.execute("DELETE FROM alimentos WHERE id = ?", (alimento_id,))
    conn.commit()
    conn.close()


def criar_grupo(nome_grupo: str):
    """Cria um novo grupo na tabela grupos."""
    nome_grupo = nome_grupo.strip()
    if not nome_grupo:
        raise ValueError("O nome do grupo não pode estar vazio.")
    
    # Verifica se já existe (case-insensitive)
    grupos_existentes = listar_grupos()
    if nome_grupo.lower() in [g.lower() for g in grupos_existentes]:
        raise ValueError(f"O grupo '{nome_grupo}' já existe.")
    
    # Insere na tabela grupos
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO grupos (nome, fonte) VALUES (?, 'usuario')", (nome_grupo,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()
    
    return nome_grupo


def excluir_grupo(nome_grupo: str):
    """Exclui um grupo da tabela grupos se não tiver alimentos associados."""
    conn = get_connection()
    cursor = conn.cursor()
    
    # Verifica se o grupo existe na tabela grupos
    grupo_row = cursor.execute("SELECT fonte FROM grupos WHERE nome = ?", (nome_grupo,)).fetchone()
    
    if not grupo_row:
        # Grupo só existe em alimentos (TACO), não pode excluir
        conn.close()
        raise ValueError(f"O grupo '{nome_grupo}' é da base TACO e não pode ser excluído.")
    
    if grupo_row["fonte"] != "usuario":
        conn.close()
        raise ValueError(f"O grupo '{nome_grupo}' é do sistema e não pode ser excluído.")
    
    # Verifica se há alimentos no grupo
    count = cursor.execute(
        "SELECT COUNT(*) as total FROM alimentos WHERE grupo = ?", 
        (nome_grupo,)
    ).fetchone()["total"]
    
    if count > 0:
        conn.close()
        raise ValueError(
            f"Não é possível excluir o grupo '{nome_grupo}' pois existem {count} "
            f"alimento(s) associado(s). Remova ou reclassifique os alimentos primeiro."
        )
    
    # Remove da tabela grupos
    cursor.execute("DELETE FROM grupos WHERE nome = ?", (nome_grupo,))
    conn.commit()
    conn.close()
    return True


def contar_alimentos_por_grupo(nome_grupo: str) -> dict:
    """Retorna contagem de alimentos de um grupo separado por fonte."""
    conn = get_connection()
    cursor = conn.cursor()
    
    total = cursor.execute(
        "SELECT COUNT(*) as n FROM alimentos WHERE grupo = ?",
        (nome_grupo,)
    ).fetchone()["n"]
    
    taco = cursor.execute(
        "SELECT COUNT(*) as n FROM alimentos WHERE grupo = ? AND fonte = 'taco'",
        (nome_grupo,)
    ).fetchone()["n"]
    
    usuario = total - taco
    
    conn.close()
    return {"total": total, "taco": taco, "usuario": usuario}


# ── Exportar / Importar ──────────────────────────────────────────────────────

def gerar_template_importacao(caminho_arquivo: str):
    """Gera planilha Excel template com cabeçalhos corretos e linhas de exemplo."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alimentos"

    colunas = ["Grupo", "Descrição", "Calorias (kcal)", "Proteínas (g)", "Carboidratos (g)", "Lipídeos (g)"]
    larguras = [28, 42, 16, 15, 17, 15]

    for col_idx, (nome, largura) in enumerate(zip(colunas, larguras), 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = largura

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    exemplos = [
        ["Cereais e derivados",      "Arroz, tipo 1, cozido",    128.0, 2.5, 28.1, 0.2],
        ["Leguminosas e derivados",  "Feijão carioca, cozido",    76.7, 4.8, 13.6, 0.5],
        ["Carnes e derivados",       "Frango, peito, grelhado",  159.0, 32.0, 0.0, 3.2],
    ]
    for row_idx, row_data in enumerate(exemplos, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(caminho_arquivo)


def exportar_alimentos_para_excel(caminho_arquivo: str, apenas_usuario: bool = False) -> int:
    """Exporta alimentos para Excel. Se apenas_usuario=True, exporta só fonte='usuario'."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_connection()
    if apenas_usuario:
        query = "SELECT grupo, descricao, calorias, proteinas, carboidratos, lipideos FROM alimentos WHERE fonte = 'usuario' ORDER BY grupo, descricao"
    else:
        query = "SELECT grupo, descricao, calorias, proteinas, carboidratos, lipideos FROM alimentos ORDER BY grupo, descricao"

    rows = conn.execute(query).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alimentos"

    headers = ["Grupo", "Descrição", "Calorias (kcal)", "Proteínas (g)", "Carboidratos (g)", "Lipídeos (g)"]
    larguras = [28, 42, 16, 15, 17, 15]

    for col_idx, (nome, largura) in enumerate(zip(headers, larguras), 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = largura

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["grupo"])
        ws.cell(row=row_idx, column=2, value=row["descricao"])
        ws.cell(row=row_idx, column=3, value=row["calorias"])
        ws.cell(row=row_idx, column=4, value=row["proteinas"])
        ws.cell(row=row_idx, column=5, value=row["carboidratos"])
        ws.cell(row=row_idx, column=6, value=row["lipideos"])

    wb.save(caminho_arquivo)
    return len(rows)


def importar_alimentos_de_excel(caminho_arquivo: str) -> dict:
    """
    Importa alimentos de um arquivo Excel.
    Valida a fonte (TACO vs usuário) comparando com a TACO embutida.
    Retorna dict com estatísticas: {criados, atualizados, erros, detalhes_erros}
    """
    import openpyxl
    import os
    import sys

    try:
        wb = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Erro ao ler arquivo Excel: {str(e)}")

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    colunas_esperadas = ['Grupo', 'Descrição', 'Calorias (kcal)', 'Proteínas (g)', 'Carboidratos (g)', 'Lipídeos (g)']
    for col in colunas_esperadas:
        if col not in headers:
            raise ValueError(f"Coluna '{col}' não encontrada no arquivo Excel.")

    col_idx = {nome: headers.index(nome) for nome in colunas_esperadas}

    # Carrega TACO embutida para validação de fonte
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    caminho_taco = os.path.join(base_path, "taco", "Taco.xlsx")
    alimentos_taco = set()
    try:
        wb_taco = openpyxl.load_workbook(caminho_taco, read_only=True, data_only=True)
        ws_taco = wb_taco.active
        taco_headers = [cell.value for cell in next(ws_taco.iter_rows(min_row=1, max_row=1))]
        gi = taco_headers.index("Grupo") if "Grupo" in taco_headers else None
        di = taco_headers.index("Descrição do Alimento") if "Descrição do Alimento" in taco_headers else None
        if gi is not None and di is not None:
            for trow in ws_taco.iter_rows(min_row=2, values_only=True):
                g = str(trow[gi] or "").strip().lower()
                d = str(trow[di] or "").strip().lower()
                if g and d and g != 'none' and d != 'none':
                    alimentos_taco.add((g, d))
        wb_taco.close()
    except Exception:
        pass

    conn = get_connection()
    cursor = conn.cursor()
    grupos_existentes = set(listar_grupos())

    criados = 0
    atualizados = 0
    erros = 0
    detalhes_erros = []

    for linha, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            grupo = str(row[col_idx['Grupo']] or "").strip()
            descricao = str(row[col_idx['Descrição']] or "").strip()

            if not grupo or grupo == 'None':
                raise ValueError("Grupo vazio")
            if not descricao or descricao == 'None':
                raise ValueError("Descrição vazia")

            calorias = float(row[col_idx['Calorias (kcal)']] or 0)
            proteinas = float(row[col_idx['Proteínas (g)']] or 0)
            carboidratos = float(row[col_idx['Carboidratos (g)']] or 0)
            lipideos = float(row[col_idx['Lipídeos (g)']] or 0)

            if any(v < 0 for v in [calorias, proteinas, carboidratos, lipideos]):
                raise ValueError("Valores negativos")

            if proteinas + carboidratos + lipideos > 100:
                raise ValueError(
                    f"Soma de macros ({proteinas + carboidratos + lipideos:.1f}g) excede 100g por 100g de alimento"
                )

            chave = (grupo.lower(), descricao.lower())
            fonte = 'taco' if chave in alimentos_taco else 'usuario'

            if grupo not in grupos_existentes:
                try:
                    criar_grupo(grupo)
                    grupos_existentes.add(grupo)
                except (ValueError, Exception) as e:
                    if 'já existe' not in str(e).lower():
                        raise

            existe = cursor.execute("""
                SELECT id, fonte FROM alimentos
                WHERE LOWER(grupo) = LOWER(?) AND LOWER(descricao) = LOWER(?)
            """, (grupo, descricao)).fetchone()

            if existe:
                cursor.execute("""
                    UPDATE alimentos
                    SET calorias = ?, proteinas = ?, lipideos = ?, carboidratos = ?
                    WHERE id = ?
                """, (calorias, proteinas, lipideos, carboidratos, existe['id']))
                atualizados += 1
            else:
                cursor.execute("""
                    INSERT INTO alimentos (grupo, descricao, calorias, proteinas, lipideos, carboidratos, fonte)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (grupo, descricao, calorias, proteinas, lipideos, carboidratos, fonte))
                criados += 1

        except Exception as e:
            erros += 1
            detalhes_erros.append(f"Linha {linha}: {str(e)}")

    wb.close()
    conn.commit()
    conn.close()

    return {'criados': criados, 'atualizados': atualizados, 'erros': erros, 'detalhes_erros': detalhes_erros}


def restaurar_base_taco() -> int:
    """
    Restaura alimentos da base TACO aos valores originais usando TACO embutida.
    Sobrescreve apenas alimentos com fonte='taco'.
    Retorna quantidade de alimentos restaurados.
    """
    import openpyxl
    import os
    import sys

    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    caminho_taco = os.path.join(base_path, "taco", "Taco.xlsx")

    if not os.path.exists(caminho_taco):
        raise ValueError(f"Arquivo TACO não encontrado em: {caminho_taco}")

    try:
        wb = openpyxl.load_workbook(caminho_taco, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Erro ao ler arquivo TACO: {str(e)}")

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    nomes_colunas = {
        "grupo": "Grupo",
        "descricao": "Descrição do Alimento",
        "calorias": "Energia(kcal)",
        "proteinas": "Proteína(g)",
        "lipideos": "Lipídeos(g)",
        "carboidratos": "Carboidrato(g)",
    }
    col_map = {}
    for campo, nome in nomes_colunas.items():
        if nome not in headers:
            raise ValueError(f"Coluna obrigatória não encontrada no arquivo TACO: '{nome}'")
        col_map[campo] = headers.index(nome)

    def _safe_float(v):
        try:
            return round(float(v or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    conn = get_connection()
    cursor = conn.cursor()
    restaurados = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        descricao = str(row[col_map["descricao"]] or "").strip()
        if not descricao or descricao == "None":
            continue

        grupo = str(row[col_map["grupo"]] or "Sem grupo").strip()
        calorias = _safe_float(row[col_map["calorias"]])
        proteinas = _safe_float(row[col_map["proteinas"]])
        lipideos = _safe_float(row[col_map["lipideos"]])
        carboidratos = _safe_float(row[col_map["carboidratos"]])

        existe = cursor.execute("""
            SELECT id FROM alimentos
            WHERE LOWER(grupo) = LOWER(?) AND LOWER(descricao) = LOWER(?) AND fonte = 'taco'
        """, (grupo, descricao)).fetchone()

        if existe:
            cursor.execute("""
                UPDATE alimentos
                SET calorias = ?, proteinas = ?, lipideos = ?, carboidratos = ?
                WHERE id = ?
            """, (calorias, proteinas, lipideos, carboidratos, existe['id']))
        else:
            cursor.execute("""
                INSERT INTO alimentos (grupo, descricao, calorias, proteinas, lipideos, carboidratos, fonte)
                VALUES (?, ?, ?, ?, ?, ?, 'taco')
            """, (grupo, descricao, calorias, proteinas, lipideos, carboidratos))

        restaurados += 1

    wb.close()
    conn.commit()
    conn.close()

    return restaurados


# ── Exportar / Importar Cardápio ─────────────────────────────────────────────

def exportar_cardapio(cardapio_id: int, caminho_arquivo: str) -> str:
    """Exporta cardápio para JSON com dados nutricionais inline. Retorna nome do cardápio."""
    import json

    cardapio = carregar_cardapio(cardapio_id)
    if not cardapio:
        raise ValueError(f"Cardápio {cardapio_id} não encontrado.")

    dados = {
        "versao": "1.0",
        "cardapio": {
            "nome": cardapio.nome,
            "kcal_total": cardapio.kcal_total,
            "data_criacao": cardapio.data_criacao,
            "refeicoes": [],
        },
    }

    for ref in cardapio.refeicoes:
        ref_dict = {
            "nome": ref.nome,
            "percentual": ref.percentual,
            "ordem": ref.ordem,
            "itens": [],
        }
        for item in ref.itens:
            ref_dict["itens"].append({
                "descricao": item.alimento.descricao,
                "grupo": item.alimento.grupo,
                "quantidade_g": item.quantidade_g,
                "calorias": item.alimento.calorias,
                "proteinas": item.alimento.proteinas,
                "carboidratos": item.alimento.carboidratos,
                "lipideos": item.alimento.lipideos,
                "fonte": item.alimento.fonte,
            })
        dados["cardapio"]["refeicoes"].append(ref_dict)

    with open(caminho_arquivo, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    return cardapio.nome


def importar_cardapio(caminho_arquivo: str) -> dict:
    """
    Importa cardápio de JSON.
    Para cada item, tenta resolver no banco local por nome+grupo (sem acento).
    Itens não encontrados são salvos como custom (dados inline).
    Retorna dict: {cardapio_id, nome, resolvidos, custom}
    """
    import json

    try:
        with open(caminho_arquivo, "r", encoding="utf-8") as f:
            dados = json.load(f)
    except Exception as e:
        raise ValueError(f"Erro ao ler arquivo: {e}")

    if "versao" not in dados or "cardapio" not in dados:
        raise ValueError("Formato inválido: arquivo não é um cardápio exportado pelo sistema.")

    raw = dados["cardapio"]

    for campo in ("nome", "kcal_total", "refeicoes"):
        if campo not in raw:
            raise ValueError(f"Arquivo inválido: campo '{campo}' ausente.")

    conn = get_connection()
    conn.create_function("sem_acento", 1, _sem_acento)

    resolvidos = 0
    custom = 0
    refeicoes = []

    for i, ref_raw in enumerate(raw["refeicoes"]):
        if "nome" not in ref_raw or "percentual" not in ref_raw:
            raise ValueError(f"Refeição {i+1} inválida: campos 'nome' ou 'percentual' ausentes.")

        itens = []
        for j, item_raw in enumerate(ref_raw.get("itens", [])):
            if "descricao" not in item_raw or "quantidade_g" not in item_raw:
                raise ValueError(f"Item {j+1} da refeição '{ref_raw['nome']}' inválido.")

            descricao = item_raw["descricao"]
            grupo = item_raw.get("grupo", "")

            # Tenta resolver no banco local (sem acento, case-insensitive)
            row = conn.execute("""
                SELECT * FROM alimentos
                WHERE sem_acento(descricao) = sem_acento(?)
                  AND sem_acento(grupo) = sem_acento(?)
                LIMIT 1
            """, (descricao, grupo)).fetchone()

            if row:
                alimento = Alimento(
                    id=row["id"],
                    grupo=row["grupo"],
                    descricao=row["descricao"],
                    calorias=row["calorias"],
                    proteinas=row["proteinas"],
                    lipideos=row["lipideos"],
                    carboidratos=row["carboidratos"],
                    fonte=row["fonte"],
                )
                resolvidos += 1
            else:
                alimento = Alimento(
                    id=None,
                    grupo=grupo,
                    descricao=descricao,
                    calorias=float(item_raw.get("calorias", 0)),
                    proteinas=float(item_raw.get("proteinas", 0)),
                    lipideos=float(item_raw.get("lipideos", 0)),
                    carboidratos=float(item_raw.get("carboidratos", 0)),
                    fonte="custom",
                )
                custom += 1

            itens.append(ItemRefeicao(alimento=alimento, quantidade_g=float(item_raw["quantidade_g"])))

        refeicoes.append(Refeicao(
            nome=ref_raw["nome"],
            percentual=float(ref_raw["percentual"]),
            kcal_diaria=float(raw["kcal_total"]),
            itens=itens,
            ordem=int(ref_raw.get("ordem", i)),
        ))

    conn.close()

    cardapio = Cardapio(nome=raw["nome"], kcal_total=float(raw["kcal_total"]), refeicoes=refeicoes)
    cardapio_id = salvar_cardapio(cardapio)

    return {"cardapio_id": cardapio_id, "nome": raw["nome"], "resolvidos": resolvidos, "custom": custom}
